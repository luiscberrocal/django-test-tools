import pytz
from datetime import datetime
import os

import subprocess
from django.conf import settings
from django.core.management import BaseCommand

import logging
from openpyxl import Workbook

from ...file_utils import add_date
from ...utils import datetime_to_local_time

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    """
    This command will export your git commits into a simple excel sheet. The report includes hash, email, date and
    description.

    Instead of using:

        git log --date=iso --pretty="format:%h|%ae|%ai|%s" > ./output/commits.txt

    You can type

        $ python manage.py git_report -f /path/to/excel.xlsx

    Note: usefulll datetime format in excel dd-mmm-yy h:mm:ss

    """

    def add_arguments(self, parser):
        # parser.add_argument('requirement_filename')
        # parser.add_argument("-l", "--list",
        #                     action='store_true',
        #                     dest="list",
        #                     help="List data",
        #                     )

        parser.add_argument("--format",
                            dest="format",
                            help="Git format",
                            default=None,
                            )

        parser.add_argument('-f', "--filename",
                            dest="filename",
                            help="Output filename",
                            default=None,
                            )

        # parser.add_argument("-u", "--username",
        #                 dest="usernames",
        #                 help="LDAP usernames for employees",
        #                 nargs='+',
        #                 )

    def handle(self, *args, **options):
        headers = ['Hash', 'Email', 'Date', 'Description']
        if options.get('format') is None:
            git_format = '%h|%ae|%ai|%s'
        else:
            git_format = options.get('format')

        if options.get('filename') is None:
            filepath = os.path.join(settings.TEST_OUTPUT_PATH, 'git_report.xlsx')
            filename = add_date(filepath)
        else:
            filename = options.get('filename')

        wb = Workbook()
        sheet = wb.create_sheet()
        row = 1
        col = 1
        for header in headers:
            sheet.cell(row=row, column=col, value=header)
            col += 1
        row += 1
        git_lines = self.report(git_format)
        for line in git_lines:
            self.stdout.write(line)
            line_data = line.split('|')
            col = 1
            for data in line_data:
                if col == 3:
                    data = self.parse_date(data)
                sheet.cell(row=row, column=col, value=data)
                col += 1
            row += 1

        wb.save(filename)

        self.stdout.write('Finished processing {}'.format(filename))

    def report(self, git_format):
        # git log --pretty=format:"%h - %an, %ad : %s" --date=iso
        #
        # git log --pretty="%h - %s" --author=gitster --since="2008-10-01"
        # --before="2008-11-01" --no-merges -- t/
        try:
            # git-describe doesn't update the git-index, so we do that
            # subprocess.check_output(["git", "update-index", "--refresh"])

            # get info about the latest tag in git
            git_commands = ["git", "log", '--date=iso', '--pretty=format:"{}"'.format(git_format)]

            describe_out = subprocess.check_output(
                git_commands,
            ).decode('utf-8')
        except subprocess.CalledProcessError:
            logger.warning("Error when running git describe")
            return {}

        return self.cleanup(describe_out.split('\n'))

    def cleanup(self, git_lines):
        cleaned_up = list()
        for git_line in git_lines:
            cleaned_up.append(git_line[1:-1])
        return cleaned_up


    def parse_date(self, date_value):
        """
        Will transform a string in the format '2018-09-23 09:37:50 -0500' to a datetime object
        :param date_value: <str> date in iso format
        :return: <datetime> datetime object of the string value
        """
        date_format = '%Y-%m-%d %H:%M:%S %z'

        datetime_object = datetime.strptime(date_value, date_format)
        datetime_object = datetime_object.replace(tzinfo=pytz.UTC)
        time_zone = pytz.timezone(settings.TIME_ZONE)

        datetime_object_with_timezone = datetime_object.astimezone(time_zone)
        return datetime_object
