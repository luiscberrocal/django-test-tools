from openpyxl import load_workbook


class ExcelAdapter(object):

    @classmethod
    def convert_to_list(cls, filename, sheet_name=None, has_header=True):
        """
        Reads an Excel file and converts every row into a list of values.

        Args:
            filename (str): Excel filename

        Kwargs:
            sheet_name (str): Name of the sheet
            has_header (boolean): If true the first row is not included in the list.

        Returns:
            list. A list of lists containg the content of the sheet


        """
        data = list()
        wb = load_workbook(filename=filename, data_only=True)
        if sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[sheet_name]
        row_num = 1
        for row in sheet.rows:
            row_data = list()
            if row_num == 1 and has_header:
                row_num += 1
                continue
            for column in range(0, len(row)):
                row_data.append(row[column].value)
            data.append(row_data)
            row_num += 1
        return data

    @classmethod
    def convert_to_dict(cls, filename, sheet_name=None):
        """
        Reads an Excel file and converts every row into a dictionary. All values are converted to strings.
        Assumes first row contains the name of the attributes.
        :param filename: <str> Excel filename
        :param sheet_name: <str> Name of the sheet
        :return: <list> A list of dictionaries.
        """
        data = list()
        wb = load_workbook(filename=filename, data_only=True)
        if sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[sheet_name]
        row_num = 1
        attributes = list()
        for row in sheet.rows:
            row_data = dict()
            if row_num == 1:
                row_num += 1
                for column in range(0, len(row)):
                    attributes.append(row[column].value)
                continue
            for column in range(0, len(row)):
                row_data[attributes[column]] = str(row[column].value)
            data.append(row_data)
            row_num += 1
        return data
